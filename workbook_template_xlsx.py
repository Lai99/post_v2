#-------------------------------------------------------------------------------
# Name:        workook_template_xlsx
# Purpose:     load .csv data and post to *.xlsx. Use openpyxl to read and write
#
# Author:      Lai
#
# Created:     12/15/2015
#-------------------------------------------------------------------------------
from openpyxl import load_workbook
from data_manage import Data

class _Abstract_Workbook_Template:
    def __init__(self, path):
        # save template sheet setup to avoid setup re-construct. This is for speed up
        self._sheet_setup= {}
        self._wb = load_workbook(path, data_only=True)
        # save explicit (name:sheet_pos) reflection to which will be use in post
        self._sheet_arrange = self._get_sheet_arrange(self._wb)

    def save(self,path):
        pass

    def post(self,data_path, mode, band, standard_anchor, channel_anchor):
        pass

    def _get_sheet_arrange(self,workbook):
        pass

    def _get_items_pos(self,sheet,anchor,items):
        pass

    def _get_fill_pos(self, mode, band, anchor):
        pass

class _Get_Sheet_Arrange:
    def __init__(self):
        self._standard_manage_func = {"2G":self._manage_standard_2G,
                                      "5G":self._manage_standard_5G
                                     }

    def _make_module_item_key(self,sheet, pos, offset):
        """
        Add modulation and rate to a string
        """
        if sheet.cell(row=pos[0],column=offset).value:
            rate = str(sheet.cell(row=pos[0],column=offset).value)

            if "." in rate:
                rate = rate.replace(".","_")

##            #if the rate from sheet is float
##            if rate.split(".")[1] == '0':
##                rate = rate.split(".")[0]
##            else:
##                rate = rate.replace(".","_")

            return self._manage_modulation(sheet,pos) + "-" + rate
        else:
            return self._manage_modulation(sheet,pos)

    def _manage_modulation(self,sheet,pos):
        """
        Draw "modulation" express from template
        """
    ##************************************************
    ## for match 2.4G 11b "DSSS-CCK". data will get "CCK" instead of "DSSS"
        if "CCK" in sheet.cell(row=pos[0],column=pos[1]).value:
            modulation = sheet.cell(row=pos[0],column=pos[1]).value.split("-")[1]
    ##************************************************
        else:
            modulation = sheet.cell(row=pos[0],column=pos[1]).value.split("-")[0]
        return modulation.strip()

    def _manage_standard_5G(self,sheet,pos):
        """
        Draw "5G standard" express from template
        """
        s = str(sheet.cell(row=pos[0],column=pos[1]).value)
        if "\n" in s:
            standard_rate, stream = s.split("\n")
            standard_rate = standard_rate.replace(" ","")
            standard, rate = standard_rate.split("-")
    ##        print standard_rate
            stream = stream.split(" ")[0]
            rate = rate.split("T")[-1]
    ##        print (standard,rate,stream)
            return (standard,rate,stream)
        else:
    ##        print s
    ##************************************************
    ## for match data "11a" actually got "11ag
            if s == "11a":
                s = "11ag"
    ##************************************************
            return s

    def _manage_standard_2G(self,sheet,pos):
        """
        Draw "2.4G standard" express from template
        """
        s = str(sheet.cell(row=pos[0],column=pos[1]).value)
        if "\n" in s:
            standard_rate, stream = s.split("\n")
            standard_rate = standard_rate.strip()
            standard, rate = standard_rate.split(" ")
            stream = stream.split(" ")[0]
            rate = rate.split("M")[0]
    ##        print (standard,rate,stream)
    ##************************************************
    ## from sheet will get "11gac" but data is "11ac"
            if standard == "11gac":
                standard = "11ac"
    ##************************************************
            return (standard,rate,stream)
        else:
    ##        print s
    ##************************************************
    ## for match data "11g" actually got "11ag
            if s == "11g":
                s = "11ag"
    ##************************************************
            return s

class _Post_Func:
    def __init__(self):
        # if data item name changed, it can modify the value reflection
        self.item_ref = {"standard":"standard",
                    "rate":"rate",
                    "channel":"channel",
                    "stream":"stream",
                    "BW":"BW"
                   }

        # key: item name in sheet, value: item name in data
        self.sheet_item_ref = {"Tx Power":"power",
                          "EVM":"EVM",
                          "Mask":"Mask",
                          "Freq error":"F_ER",
                          "SC error":"CR_ER",
                          "Flatness":"Flatness",
                          "Rx Power":"SENS"
                         }

        self.rx_item = ["SENS"]

        self.post_func= {# TX
                        "power":self._post_power,
                        "EVM":self._post_evm,
                        "Mask":self._post_mask,
                        "F_ER":self._post_freq_err,
                        "CR_ER":self._post_cr_err,
                        "Flatness":self._post_flatness,
                        # RX
                        "SENS":self._post_sens
                        }

    def _post_power(self,data):
        if "power" in data:
            if data["power"]:
                return data["power"].split(",")
            return data["power"]
        return None

    def _post_evm(self,data):
        if "EVM" in data:
            if data["EVM"]:
                return data["EVM"].split(",")
            return data["EVM"]
        return None

    def _post_mask(self,data):
        if "Mask" in data:
            if data["Mask"]:
                return data["Mask"].split(",")
            return data["Mask"]
        return None

    def _post_freq_err(self,data):
        if "F_ER" in data:
            if data["F_ER"]:
                return list([data["F_ER"]])
            return data["F_ER"]
        return None

    def _post_cr_err(self,data):
        if "CR_ER" in data:
            if data["CR_ER"]:
                return list([data["CR_ER"]])
            return data["CR_ER"]
        return None

    def _post_flatness(self,data):
        if "Flatness" in data:
            if data["Flatness"]:
                return data["Flatness"].split(":")
            return data["Flatness"]
        return None

    def _post_sens(self,data):
        if "sens" in data:
            if data["sens"]:
                return data["sens"].split(",")
            return data["sens"]
        return None

    def _check_same_row(self,data, last_data_conf):
        """
        Check if the same sheet row. It will check that "standard", "rate", "band width", and "stream" are all meet
        """
        if not last_data_conf:
            return False

        conf = self._get_data_conf(data)
        if conf == last_data_conf:
            return True
        return False

    def _get_data_conf(self,data):
        return (data["standard"],data["rate"],data["BW"],data["stream"])

    def _post_value(self,ws,data,start,ch_pos,case_num):
        """
        Post value in sheet explicit position
        """
        for i in range(case_num):
            # Get item name
            case = ws.cell(row=start[0]+i,column=start[1]-1).value
    ##        print case
            # if valid item name
            if case in self.sheet_item_ref:
                value = self.post_func[self.sheet_item_ref[case]](data)
    ##            print data
    ##            print value
                antennas = data["antenna"].split(",")
    ##            print antennas, len(antennas)
                if value:
                    # value > 1 means multiple streams
                    if len(value) > 1:
                        for idx in range(int(data["stream"])):
    ##                    for idx in range(len(antennas)):
                            post_pos = (start[0]+i,ch_pos[1]+int(antennas[idx]))
##                            print type(value[idx]), value[idx]
                            ws.cell(row=post_pos[0],column=post_pos[1]).value = value[idx]
                    else:
                        # antennas = 1 and value = 1 will be 1 stream
                        if len(antennas) > 1:
    ###################### for RX 11ac MIMO and SIMO post #######################
                            if self.sheet_item_ref[case] in self.rx_item:
                                move = self._find_ch_sum(ws,ch_pos)
    ##                            print move
                                post_pos = (start[0]+i,ch_pos[1]+move)
    #############################################################################
                            else:
                                # TX no need move
                                post_pos = (start[0]+i,ch_pos[1])
                        else:
                            post_pos = (start[0]+i,ch_pos[1]+int(antennas[0]))
##                        print type(value[0]), value[0]
                        ws.cell(row=post_pos[0],column=post_pos[1]).value = value[0]

    def _meet_standard(self,data,fill_pos):
        """
        Get "standard" in sheet position
        """
    ############ RX 11n need this to let "stream" to meet really config ###################
    ############ MIMO will always get stream '1'
        if data[self.item_ref["standard"]] == "11n":
            ch = int(data[self.item_ref["rate"]][3:])
            if ch < 8:
                stream = '1'
            elif ch < 16:
                stream = '2'
            elif ch < 24:
                stream = '3'
            else:
                stream = '4'
            k = (data[self.item_ref["standard"]], data[self.item_ref["BW"]], stream)
    #######################################################################################
        else:
            k = (data[self.item_ref["standard"]], data[self.item_ref["BW"]], data[self.item_ref["stream"]])
    ##    print k
    ##    print fill_pos.keys()
        #
        if k in fill_pos:
            return fill_pos[k]

        # MCSx
        if k[0] in fill_pos:
            return fill_pos[k[0]]
        print "Can't find this channel " + data[self.item_ref["channel"]] + "," + data[self.item_ref["standard"]]
        return None

    def _meet_rate(self,data,fill_pos):
        """
        Get "rate" in sheet position
        """
    ##    print fill_pos.keys()
    ##    print data[item_ref["rate"]]
        for k in fill_pos.keys():
            if "-" in data[self.item_ref["rate"]]:
                if data[self.item_ref["rate"]] == k:
                    return fill_pos[k]
            else:
                if "-" in k:
                    #  data[item_ref["rate"]] with modulation
                    if data[self.item_ref["rate"]] == k.split("-")[0]:
                        return fill_pos[k]
                    # data[item_ref["rate"]] only have rate
                    elif data[self.item_ref["rate"]] == k.split("-")[1]:
                        return fill_pos[k]
                else:
                    if data[self.item_ref["rate"]] == k:
                        return fill_pos[k]
        print "Can't find this modulation " + data[self.item_ref["rate"]] + "," + data[self.item_ref["standard"]]
        return None

class Workbook_Template_Xlsx(_Abstract_Workbook_Template, _Get_Sheet_Arrange, _Post_Func):
    def __init__(self,path):
        _Abstract_Workbook_Template.__init__(self,path)
        _Get_Sheet_Arrange.__init__(self)
        _Post_Func.__init__(self)

    def save(self,path):
        self._wb.save(path + ".xlsx")

    def post(self,data_path, mode, band, standard_anchor, channel_anchor):
        """
        Get template setup to find post position then post value
        """
        ws = self._wb.worksheets[self._sheet_arrange[mode+band]]
        self._max_col = ws.max_column
        self._max_row = ws.max_row

        if not (mode+band) in self._sheet_setup:
            self._sheet_setup[mode+band] = self._get_fill_pos(ws, mode,band,standard_anchor)

##        for i in self._sheet_setup[mode+band][0]:
##            for j in self._sheet_setup[mode+band][0][i]:
##                print i,j,self._sheet_setup[mode+band][0][i][j]

        fill_pos, all_anchor_row = self._sheet_setup[mode+band][0], self._sheet_setup[mode+band][1]

        last_data_conf = None
        need_pos = None
        case_num = 0
        ch_start = None
        ch_pos = None
        ch_now = None
        last_ch = 0

        input_data = Data(data_path)

        for data in input_data.load_data():
            if not self._check_same_row(data, last_data_conf):
    ##            print data
    ##            print fill_pos.keys()
                # find "standard" position
                standard_pos = self._meet_standard(data, fill_pos)
                if not standard_pos:
                    continue

                # find "rate" position
                rate_pos = self._meet_rate(data, standard_pos)
                if rate_pos:
                    need_pos, case_num = rate_pos
                    last_data_conf = self._get_data_conf(data)
                else:
                    continue
    ##            print need_pos, case_num
                # Get post start position
                ch_start = self._get_channel_start(need_pos,all_anchor_row)
                ch_now = ch_start
    ##        print ch_start, "ch start"
            # Get value post position
    ##            print data[item_ref["channel"]]
            if int(data[self.item_ref["channel"]]) > last_ch:
                ch_pos = self._get_channel_pos(ws,ch_now,data[self.item_ref["channel"]])
            else:
                ch_pos = self._get_channel_pos(ws,ch_start,data[self.item_ref["channel"]])
            last_ch = int(data[self.item_ref["channel"]])
##            print ch_pos, "ch pos"
            if ch_pos:
                self._post_value(ws,data,need_pos,ch_pos,case_num)
                # if value appear in ch by ch will lose
                ch_now = ch_pos
            else:
                continue

    def _get_sheet_arrange(self,workbook):
        """
        To find sheet name "TX / RX" and "2.4G / 5G" and make a dict (name:sheet_pos)
        """
        sheet_names = [i.lower() for i in workbook.get_sheet_names()]
        sheet_ref = {}
        for idx in range(len(sheet_names)):
            if "2.4ghz" in sheet_names[idx]:
                if "tx" in sheet_names[idx]:
                    sheet_ref["TX2G"] = idx
                elif "sensitivity" in sheet_names[idx]:
                    sheet_ref["RX2G"] = idx
            elif "5ghz" in sheet_names[idx]:
                if "tx" in sheet_names[idx]:
                    sheet_ref["TX5G"] = idx
                elif "sensitivity" in sheet_names[idx]:
                    sheet_ref["RX5G"] = idx
        return sheet_ref

    def _get_items_pos(self,mode):
        """
        Recall template setup if it exist. If template setup isn't exist, call "get_fill_pos" to make.
        """
        if mode == "TX":
            # standard_x = 1,module_x = 2,rate_x = 3, case_x = 5, start_x = 6
            return 1,2,3,5,6
        else:
            # standard_x = 1,module_x = 2,rate_x = 3, case_x = 6, start_x = 7
            return 1,2,3,6,7

    def _get_fill_pos(self, sheet, mode, band, anchor):
        """
        Get all value can be filled position in a sheet
        Input: int:specified sheet, string:anchor which used to split data block, int:band
        Output:dict:key:whole sheet value can be filled position, value:all anchors row loocation
        """
        standard_x ,module_x ,rate_x , case_x, start_x = self._get_items_pos(mode)

        start = 0
        all_anchor_row = []
        #Don't need sheet front content. Use anchor to go to standard start position
        for row in range(1,50):
            if sheet.cell(row=row,column=standard_x).value == anchor:
                start = row
                all_anchor_row.append(row)
                break

        if len(all_anchor_row) == 0:
            "Find no sheet anchor"
            return 1
##        print all_anchor_row

        last_standard = (0,0)
        last_module = (0,0)
        items = {}
        module_items = {}
        case_count = 0

        for row in range(start + 1,self._max_row + 1):
            #Use standard between standard to split data block, need to add last_standard one in the end
            # Warning! Might include none test item count because no clue to check it is item or not
            if sheet.cell(row=row,column=module_x).value != None and sheet.cell(row=row,column=module_x).value != "":    #Collect Modulations in a standard
                if case_count > 0:
                    # For eliminate the warning point out thing
                    if sheet.cell(row=row,column=standard_x).value == anchor:
                        case_count -= 1
                    #Add "module and rate" with "value start position and case numbers"
                    k = self._make_module_item_key(sheet, last_module, rate_x)
                    module_items[k] = ((last_module[0], start_x),case_count)
                    case_count = 0
                last_module = (row,module_x)

            if sheet.cell(row=row,column=standard_x).value != None and sheet.cell(row=row,column=standard_x).value != "":
                if  sheet.cell(row=row,column=standard_x).value != anchor:
                    if module_items:   #if true means it has a modulation collection
    ##************************************************************************************************
    ##  5G sheet has a sheet tail. When reach this, stop search and record "standard"
                        if sheet.cell(row=row,column=standard_x).value == "Info":
                            break
    ##************************************************************************************************
                        items[self._standard_manage_func[band](sheet,last_standard)] = module_items
    ##                    print module_items.values()
                        module_items = {}
                    last_standard = (row,standard_x)  #A spec start position
                else:
                    all_anchor_row.append(row)

                    continue   #Not include row which has anchor

            if sheet.cell(row=row,column=case_x).value != None and sheet.cell(row=row,column=case_x).value != "":  #Count how many test case
                case_count += 1

        #Don't forget last one have no end point
        if case_count != 0:
            #Add "module and rate" with "value start position and case numbers"
            k = self._make_module_item_key(sheet, last_module, rate_x)
            module_items[k] = ((last_module[0], start_x),case_count)

        if module_items:
            items[self._standard_manage_func[band](sheet,last_standard)] = module_items
        return items, all_anchor_row

    def _get_channel_start(self, pos, all_anchor_row = None):
        """
        Search row in all_anchor_row that closest to pos. The row have the channel information
        """
        row, col = pos[0], pos[1]
        if all_anchor_row:
            if row - all_anchor_row[0] > 0:
                if len(all_anchor_row) == 1:
                    return (all_anchor_row[0],col)
                else:
                    for i in range(len(all_anchor_row)):
                        if row - all_anchor_row[i] < 0:
                            return (all_anchor_row[i-1],col)
                    return (all_anchor_row[-1],col)
        print "Can't find channel form location in sheet"
        return None

    def _get_channel_pos(self, sheet, pos, ch):
        """
        Search column to find the channel title position
        """
        row, col = pos[0], pos[1]

        count = 31
        # beacause it might have blank, need to pass
        while count > 0:
            if col <= self._max_col:
                while sheet.cell(row=row,column=col).value:
                    if ch in sheet.cell(row=row,column=col).value:
                        return (row, col)
                    col += 1
                    count = 31
                col += 1
                count -= 1
            else:
                break
        print "Can't find this channel in channel form " + str(ch) + " , "+ str(pos)
        return None

    def _find_ch_sum(self, sheet, ch_pos):
        """
        Search column to find the last block value meet value in ch_pos
        """
        count = 0
        match = sheet.cell(row=ch_pos[0], column=ch_pos[1]).value.replace(" ","")

        while ch_pos[1]+count <= self._max_col and sheet.cell(row=ch_pos[0],column=ch_pos[1]+count).value.replace(" ","") == match:
##            print ch_pos[0] , ch_pos[1] + count
            count += 1
        return count - 1